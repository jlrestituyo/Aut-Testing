import openpyxl

def get_row_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    return sheet.max_row

def get_column_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    return sheet.max_column

def read_data_from_excel(file, sheet_name, row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row = row_num, column = column_num).value

def write_data_to_excel(file, sheet_name, row_num, column_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row = row_num, column = column_num).value = data
    workbook.save(file)


def get_data(file, sheet):
    rows = get_row_count(file, sheet)
    columns = get_column_count(file, sheet)
    fila =0
    columna = 0
    test_data = [ [0 for fila in range(2,4)] for columna in range (1,3)]
    for i in range(2, 4):
        for j in range(1, 3):
            test_data[fila].append(read_data_from_excel(file, sheet, i, 1))
            test_data[fila].append(read_data_from_excel(file, sheet, i, 2))

    return test_data